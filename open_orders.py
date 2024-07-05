import datetime
import os
import pandas as pd
from sqlalchemy.exc import ResourceClosedError
from sqlalchemy import create_engine, text
import fabory_logging
from sys import exit
import openpyxl
import xlsxwriter
from openpyxl.utils import get_column_letter
import time


# Define email address for error handling
ict_dosupport = "EmailAddress"

# Initialize a logger using fabory_logging module
log = fabory_logging.fabory_logger(__name__)

# Defines a list of column names to check
column_names_to_check = ['Column', 'Names']

# Defines a list of column names that will be colored to check
colored_column_names_to_check = ['Column', 'Names']

# Disable SettingWithCopyWarning
pd.options.mode.chained_assignment = None  # default='warn'

#current date in format dd.mm.yyyy
current_date = datetime.datetime.now().strftime("%d.%m.%Y")

# Get today's date
today_date = datetime.datetime.today()

# Get current date and time in format yyyy-mm-dd 00:00:00
formatted_date = datetime.datetime.now().strftime("%Y-%m-%d 00:00:00")

# Server connection details for the EOL_app database
_SERVER_ = ''    # Server IP or hostname
_DATABASE_ = ''  # Database name
_USERNAME_ = ''  # Database username
_PASSWORD_ = ''  # Database password
_PORT_ = ""      # Port for database communication

# Construct the connection string for the database
database_connection = f"SetDatabaseConnection"

def check_recipients_file(recipients_excel_file, copy_recipients_excel_file):
    """
    Checks if the recipients' Excel file and the copy recipients' Excel file exist.
    If they exist, reads the email addresses from the files into lists and returns them.
    If not, logs an error and exits the program.
    """

    if not os.path.exists(recipients_excel_file):

        # Log an error if recipients_excel_file file is not found
        log.error(f"Path to {recipients_excel_file} was not found.")
        exit()

    elif not os.path.exists(copy_recipients_excel_file):

        # Log an error if copy_recipients_excel_file Excel file is not found
        log.error(f"Path to {copy_recipients_excel_file} was not found.")
        exit()

    else:
        # Read recipients' email addresses from the Excel file into a list
        recipients = pd.read_excel(recipients_excel_file).values.tolist()

        # Read copy recipients' email addresses from the Excel file into a list
        copy_recipients = pd.read_excel(copy_recipients_excel_file).values.tolist()

        return recipients, copy_recipients


def get_users(users):
    """
    Extracts user emails from a list of users.


    Args:
    - users (list): List containing user information.


    Returns:
    - str: Concatenated string of user emails.
    """

    # Initialize an empty string to store user emails
    users_emails = ""

    # Iterate through the users list using enumerate to get both index and user
    for index, user in enumerate(users):

        # Extract the first element (assuming it contains user information)
        user = users[index][0]

        # Check if the extracted user is a string
        if isinstance(user, str):
            # If it's a string, append it to the users_emails string
            users_emails = users_emails + user + "; "
        else:
            # If it's not a string, skip to the next iteration
            continue

    # Remove the trailing "; " and return the final string of user emails
    return users_emails[:-2]


def set_excel_data(filtered_data, existing_column_names, server_path):
    
    """
    Writes filtered_data to an Excel file at server_path with specific formatting.


    Args:
    - filtered_data (DataFrame): The DataFrame to be written to Excel.
    - server_path (str): The path where the Excel file will be saved.
    """

    # Extract column information
    column_names = filtered_data.columns
    column_indices = range(1, len(column_names) + 1)
    column_letters = [get_column_letter(idx) for idx in column_indices]

    # Columns that exists in data frame will be formatted
    column_names_with_format = filtered_data[existing_column_names]


    # Write DataFrame to Excel file
    with pd.ExcelWriter(server_path, engine='xlsxwriter') as writer:
        filtered_data.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Set date format for specific columns
        date_format = workbook.add_format({'num_format': 'dd.mm.yyyy'})

        # Iterate through columns
        for col_num, column in enumerate(column_names, start=0):

            # Check if the column needs special formatting
            if column in column_names_with_format:
                letter = column_letters[col_num]

                # Set date format for the entire column
                worksheet.set_column(f"{letter}:{letter}", None, date_format)

                # Iterate through rows in the specified column
                for index, cell_value in enumerate(filtered_data[column]):
                    cell_reference = f'{letter}{index + 2}'  # Adding 2 because Excel row indices start from 1

                    date_value_str = cell_value

                    # Check if the cell has a valid date value
                    if pd.notna(date_value_str):

                        date_object = datetime.datetime.strptime(date_value_str, "%d.%m.%Y")


                        # Extract year, month, and day from the date
                        day_value, month_value, year_value = date_object.day, date_object.month, date_object.year

                        # Apply the DATE formula to the cell
                        formula = f'=DATE({year_value}, {month_value}, {day_value})'
                        worksheet.write_formula(cell_reference, formula)


def adjust_color_width_in_excel(server_path):

    """
    Adjusts the width of cells in an Excel file based on the max length of cells in the Excel file.

    Parameters:
    - server_path (str): The file path to the Excel file located on the server.

    Note:
    - This function modifies the input Excel file in place.

    Example Usage:
    adjust_color_width_in_excel("\\\\pricing-db01\\temp_files\\CHECKLISTS_OPENORDERS_CEE_1500_RPA_01.01.1900.xlsx")

    :param server_path: serverPath
    :return:
    """

    # Load the workbook
    wb = openpyxl.load_workbook(server_path)


    # Select the active sheet
    sheet = wb.active


    # Freeze first row in Excel file
    sheet.freeze_panes = "A2"


    # Define the fill color as yellow
    yellow_fill = openpyxl.styles.PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')


    # Define the fill color as red
    red_fill = openpyxl.styles.PatternFill(start_color='FF4d4d', end_color='FF4d4d', fill_type='solid')


    # Get the first row
    first_row = sheet[1]


    # Lists for storing column indexes based on specific criteria
    not_blank_columns = ['Column', 'Names']
    incomplete_columns = ['Column', 'Names']
    matst_columns = ['Column', 'Names']
    mrpcontr_columns = ['Column', 'Names']
    check_columns = ['Column', 'Names']


    not_blank_columns_indexes = []
    incomplete_columns_indexes = []
    matst_columns_indexes = []
    mrpcontr_columns_indexes = []
    check_columns_indexes = []

    # Iterate through cells in the first row
    for cell in first_row:

        # Get the column letter index
        column_letter_index = get_column_letter(cell.column)


        # Check criteria and store column indexes accordingly
        if cell.value in not_blank_columns:
            not_blank_columns_indexes.append(column_letter_index)


        if cell.value in incomplete_columns:
            incomplete_columns_indexes.append(column_letter_index)


        if cell.value in matst_columns:
            matst_columns_indexes.append(column_letter_index)


        if cell.value in mrpcontr_columns:
            mrpcontr_columns_indexes.append(column_letter_index)


        if cell.value in check_columns:
            check_columns_indexes.append(column_letter_index)

    # Turn on filters for all columns
    sheet.auto_filter.ref = sheet.dimensions


    # Iterate through all columns
    for column in sheet.columns:
        # Initialize max_length for each column
        max_length = 0

        # Iterate through cells in the column
        for cell in column:

            # Check if cell value should be colored and apply respective fill color
            if not cell.value in colored_column_names_to_check:

                if cell.column_letter in not_blank_columns_indexes and cell.value is not None:
                    cell.fill = red_fill


                if cell.column_letter in incomplete_columns_indexes:
                    if cell.value == 'Value':
                        cell.fill = red_fill


                if cell.column_letter in matst_columns_indexes:
                    if cell.value in ['Value1', 'Value2', 'Value3']:
                        cell.fill = red_fill


                if cell.column_letter in mrpcontr_columns_indexes:
                    if cell.value in ['Value1', 'Value2', 'Value3']:
                        cell.fill = red_fill


                if cell.column_letter in check_columns_indexes:
                    if cell.value in ['Value1', 'Value2', 'Value3']:
                        cell.fill = red_fill

            else:

                cell.fill = yellow_fill

            try:

                # Check if the length of the cell value is greater than the current max_length
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

            except:
                # Ignore errors (e.g., if the cell value is None)
                pass

        # Adjust the column width based on the max_length with some padding
        adjusted_width = (max_length + 6)  # Adding some padding
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    # Save the modified workbook
    wb.save(server_path)


def send_mail(path_to_file, recipients_emails, copy_recipients_emails, msg, connection):

    """
    Sends an email using SQL Server's sp_send_dbmail stored procedure.


    Args:
    - recipients_emails (str): Email addresses of the primary recipients.
    - copy_recipients_emails (str): Email addresses of the copy recipients.
    - msg (str): Body of the email message.
    - connection: SQLAlchemy database connection object.
    """

    # No file is being sent if the email address is ict_dospport@fabory.com
    if recipients_emails == ict_dosupport:

        path_to_file = ""

    # Construct the query to send an email using the SQL Server's sp_send_dbmail stored procedure
    query = f"""SQL_QUERY"""

    log.info(f"Email has been successfully sent to the following email address(es). Recipients: {recipients_emails}, Copy Recipients: {copy_recipients_emails}")

    try:

        # Execute the query to send the email and commit changes to the connection
        data = pd.read_sql_query(text(query), con=connection)
        connection.commit()

    except ResourceClosedError:
        # Ignore the ResourceClosedError
        connection.commit()

    except Exception as error:

        log.error(f"Error {error}")
        # In case of an exception, commit changes to the connection to ensure data consistency
        connection.commit()

def get_database_connection():

    try:
        # Create a database engine using the connection string
        engine = create_engine(database_connection)

        # Establish a connection to the database using the engine
        db_connection = engine.connect()

        return db_connection

    except Exception as error:

        # Exception handling is implemented to catch any unforeseen errors,
        # printing a message if there is a failure in the database connection.
        log.error(f"Connection to database: {database_connection} failed.")
        exit()

def main(time):

    """
    Main function to execute the script.

    Note:
    - The script is designed to connect to a database, read data from an Excel file,
    filter the data, save it to a new Excel file, adjust cell widths, and send an email.

    Example Usage:
    main()
    """

    sales_orgs = None

    # Get database connection
    connection = get_database_connection()

    try:

        if time == "9am":

            sales_orgs = ['Value1', 'Value2', 'Value3']

        elif time == "1pm":

            sales_orgs = ['Value1']

        # Iterate through each sales organization
        for sales_org in sales_orgs:

            # Define file paths for recipients and copy recipients Excel files for the current sales organization
            recipients_excel_file = f"/Path/To/File"
            copy_recipients_excel_file = f"/Path/To/File"

            # Define the path to the SAL_OPENORDERS_CEE folder and specific sales org
            folder_path = f"/Path/To/Folder"

            # Define file name
            file_name = f"FileName"

            # Construct the full file path by joining the folder path and file name
            file_path = os.path.join(folder_path, file_name)


            # Check if the file exists
            if not os.path.exists(file_path):

                # Log an error message if the file path is not found
                log.error(f"Path: {file_path} was not found.")
                msg = f"ERROR: {file_path} was not found."
                send_mail(file_name, ict_dosupport, "", msg, connection)

            else:

                # Read Excel file into a pandas DataFrame
                data_frame = pd.read_excel(file_path, dtype=str)


                # Initialize lists to store existing column names and existing colored column names
                existing_column_names = []
                existing_colored_column_names = []


                # Iterate through each column name to check
                for column_name in column_names_to_check:

                    # Check if the column name exists in the DataFrame's columns
                    if column_name in data_frame.columns:

                        # Append the existing column name to the list
                        existing_column_names.append(column_name)


                # Iterate through each colored column name to check
                for colored_column_name in colored_column_names_to_check:

                    # Check if the column name exists in the DataFrame's columns
                    if colored_column_name in data_frame.columns:

                        # Append the existing colored column name to the list
                        existing_colored_column_names.append(colored_column_name)


                # Iterate through each existing column name
                for column in existing_column_names:

                    # Convert values in the existing column to datetime objects with specified format
                    data_frame[f"{column}"] = pd.to_datetime(data_frame[f"{column}"], format='%Y-%m-%d 00:00:00', errors='coerce')

                filtered_data = data_frame

                current_day_in_month = datetime.datetime.now().day

                # Check if day in a month is less thnen 26. day in a month
                if current_day_in_month < 26:

                    # Filter the DataFrame based on conditions
                    filtered_data = filtered_data[(filtered_data['Requested'].isna()) | (filtered_data['Requested'] < formatted_date)]


                # Iterate through each existing column name
                for column in existing_column_names:

                    # Convert values in the existing column back to string format with custom date format
                    filtered_data[f"{column}"] = filtered_data[f"{column}"].dt.strftime("%d.%m.%Y")


                # Log information about data filtering process
                log.info(f"Information: {sales_org}.")

                # Define server path for Excel file
                server_path = f"/Saving/Excel/File/To/Server/Folder"

                # Set Excel data and adjust color width in Excel
                set_excel_data(filtered_data, existing_column_names, server_path)
                adjust_color_width_in_excel(server_path)


                # Check if the filtered data is empty
                if filtered_data.empty:
                    msg = f"The excel file contains no data."
                    log.info("The excel file contains no data.")
                    send_mail(server_path, ict_dosupport, "", msg, connection)


                else:
                    # Generate message and send email with filtered data attached
                    msg = f"Some text in email body."
                    recipients, copy_recipients = check_recipients_file(recipients_excel_file, copy_recipients_excel_file)
                    recipients_emails, copy_recipients_emails = get_users(recipients), get_users(copy_recipients)
                    send_mail(server_path, recipients_emails, copy_recipients_emails, msg, connection)


    except Exception as error:
        # Handle any exceptions and send error message
        msg = f"Error: {error}.".replace("'", "")
        log.error(f"Error: {error}.")
        send_mail(file_name, ict_dosupport, "", msg, connection)

def run_at_specific_times():

    counter = 0;

    while True:
        current_time = datetime.datetime.now().time()

        counter += 1

        if counter == 1:

            # Log a debug message to indicate the start of a script
            log.debug("Starting a script.")

        # Check if it's 9am
        if current_time.hour == 9 and current_time.minute == 0 and current_time.second == 30:
            main("9am")

        # Check if it's 1pm
        elif current_time.hour == 13 and current_time.minute == 5 and current_time.second == 0:
            main("1pm")
            exit()  # Exit the script after main("1pm") is finished

        if current_time.minute == 0 and current_time.second == 0:
            log.info("Script is in progress..")

        # Sleep for a second before checking again
        time.sleep(1)

if __name__ == "__main__":
    run_at_specific_times()